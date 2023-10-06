import sqlite3
import os
from tkinter import filedialog
from tkinter import messagebox as mbox 
import pandas as pd
from tkinter import * 
from openpyxl import load_workbook
import openpyxl
from datetime import date
from datetime import datetime
import sys

#Create a Window to select db and xlxs file
def selectFile(titleText, typeOfFile, extFile, textObj): 
    textObj.delete(0,"end")
    fileName = filedialog.askopenfilename(initialdir = os.getcwd(), title = titleText, filetypes=[(typeOfFile, extFile)])
    textObj.insert(INSERT, fileName)
    return fileName

def radioButtonOptions(opt):
    if opt == 0:
        mbox.showerror('Select an Option','Error. You must select an option.')
    else:
        ventana.quit()
        
def selec(opt):
    if opt == 4:
        labelDB.configure(text='Select Audit File:')
        btnDB.place_forget()
        btnDB2.place(x=410, y=20)
    else:
        btnDB2.place_forget()
        btnDB.place(x=410, y=20)
        labelDB.configure(text='Select Data Base:')

ventana = Tk() 
ventana.geometry("450x260") 
ventana.title('Parameter Audit') 

labelDB = Label(ventana, text ="Select Data Base:") 
labelDB.place(x=10, y=20)
valueTextDB = StringVar()
textDB = Entry(ventana, width=40, textvariable=valueTextDB)
textDB.place(x=160, y=20)
btnDB2 = Button(ventana, text ="...")
btnDB2.place(x=410, y=20)
btnDB2.bind("<Button>",  lambda e: selectFile('Select Audit File', 'Excel', '*.xlsx', textDB))
btnDB = Button(ventana, text ="...")
btnDB.place(x=410, y=20)
btnDB.bind("<Button>",  lambda e: selectFile('Select Data Base', 'database', '*.db', textDB))

labelAF = Label(ventana, text ="Select Audit Template:")
labelAF.place(x=10, y=60)
valueTextAF = StringVar()
textAF = Entry(ventana, width=40, textvariable=valueTextAF)
textAF.place(x=160, y=60)
btnAF = Button(ventana, text ="...")
btnAF.place(x=410, y=60)
btnAF.bind("<Button>",  lambda e: selectFile('Select Audit Template', 'Excel', '*.xlsx', textAF)) 

labelOptions = Label(ventana, text ="OPTIONS:") .place(x=10, y=100)

valueRadioBtn = IntVar()
RdoBtn = Radiobutton(ventana, text='Create Baseline (from DB) Only', variable=valueRadioBtn, value=1, command=lambda: selec(1))
RdoBtn.place(x=30, y=120)

RdoBtn = Radiobutton(ventana, text='Create Baseline (from DB) and Audit', variable=valueRadioBtn, value=2, command=lambda: selec(2))
RdoBtn.place(x=30, y=140)

RdoBtn = Radiobutton(ventana, text='Create Baseline (from DB), Audit and xml Plans', variable=valueRadioBtn, value=3, command=lambda: selec(3))
RdoBtn.place(x=30, y=160)

RdoBtn = Radiobutton(ventana, text='Create xml Plans from Audit File', variable=valueRadioBtn, value=4, command=lambda: selec(4))
RdoBtn.place(x=30, y=180)

btnStart = Button(ventana, text ="START", height=2, width=10, command=lambda: radioButtonOptions(valueRadioBtn.get()))
btnStart.place(x=200, y=210)

labelCredits = Label(ventana, text ="Design by: Fabio Urrea") .place(x=320, y=240)

ventana.mainloop()

option = valueRadioBtn.get()
routeDB = valueTextDB.get()
routeTemplate = valueTextAF.get()

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Create the Query
if option != 4 and option != 0:
    df_query = pd.read_excel(io= routeTemplate, sheet_name='Query')
    df_query.fillna("", inplace=True)

    query = "SELECT DISTINCT "
    dictJoin = {}
    arrParamTemp = []
    from_ = ''
    paramJoin = ''
    paramOrderBy = ''
    countOrderBy = 0
    for i in df_query.index.values:
        mo = df_query.iloc[i, [0]].values[0]
        param = df_query.iloc[i, [1]].values[0]
        paramTemp = df_query.iloc[i, [3]].values[0]
        concat = df_query.iloc[i, [4]].values[0]
        join = df_query.iloc[i, [5]].values[0]
        key = df_query.iloc[i, [6]].values[0]
        orderBy = df_query.iloc[i, [7]].values[0]
        parameter = mo + '.' + param

        if from_ == '':
            from_ = 'FROM ' + df_query.iloc[i, [0]].values[0] + ' '

        if concat != '':
            parameter = parameter + ' || '
            if df_query.iloc[i + 1, [4]].values[0] != concat:
                parameter = parameter + ' AS ' + "'" + concat + "'" + ', '
                parameter = parameter.replace('||  AS', 'AS')

        elif i == df_query.shape[0] - 1:
            parameter = parameter + ' '
        else:
            parameter = parameter + ', '
        
        if join != '' and mo not in dictJoin:
            dictJoin[mo] = join
            paramJoin = paramJoin + 'LEFT JOIN ' + mo + ' USING (' + join + ') '

        if paramTemp == 'NO':
            if concat != '':
                arrParamTemp.append(concat)
            else:
                arrParamTemp.append(param)
        
        if key != '':
            paramKey = key
        
        if orderBy != '':
            if countOrderBy == 0:
                paramOrderBy = paramOrderBy + 'ORDER BY ' + mo + '.' + param + ', '
                countOrderBy += 1
            else:
                paramOrderBy = paramOrderBy +  mo + '.' + param + ', '
            
        query = query + parameter

    if len(paramOrderBy) > 0:
        paramOrderBy = paramOrderBy[0:-2]

    df_queryWhere = pd.read_excel(routeTemplate, sheet_name='Where')
    paramWhere = ''
    if df_queryWhere.shape[0] > 0:
        paramWhere = 'WHERE '
        for i in df_queryWhere.index.values:
            paramWhere = (paramWhere + str(df_queryWhere.iloc[i, [0]].values[0]) + '.' + str(df_queryWhere.iloc[i, [1]].values[0]) + ' = '
                            + str(df_queryWhere.iloc[i, [2]].values[0]) + ' ' + str(df_queryWhere.iloc[i, [3]].values[0]) + ' ')

        paramWhere= paramWhere.replace(' nan', ' ')

    query = query + from_ + paramJoin + paramWhere + paramOrderBy

    # Create a Dataframe with Query information
    sql = sqlite3.connect(routeDB)
    cur = sql.cursor()
    df = pd.read_sql_query(query, sql)
    sql.close()
    templateName = routeTemplate.replace('Template', '').replace('.xlsx', '').split('_')
    templateName = '_' + templateName[-1]
    try:
        df.to_excel(os.getcwd() + '\Baseline_Query' + templateName + '.xlsx', index=False)
    except:
        mbox.showerror('File is Open','Error. File Baseline_Query is Open. Please close file and try again.')
        sys.exit()

    # Create a Dictionaries and data frame with the table of Parameters and export in xlsx files
    if option == 1:
        mbox.showinfo('Parameters to Audit', 'Excel File Baseline_Query was created')
        sys.exit()
    elif option == 2 or option == 3:
        dict_parameters = {}
        dict_parameters_object = {}
        dfParameterTemplate = pd.read_excel(routeTemplate, sheet_name='Parameter Template')
        dfParameterTemplate.fillna("", inplace=True)

        dictParamIsList = {}
        dictListWithParam = {}
        for i in dfParameterTemplate.index.values:
            key = dfParameterTemplate.iloc[i, [2]].values[0]
            valueTemplateIntValue = dfParameterTemplate.iloc[i, [4]].values[0]
            paramIsList = dfParameterTemplate.iloc[i, [5]].values[0]
            valueObject = dfParameterTemplate.iloc[i, [0]].values[0]
            dict_parameters[key] = valueTemplateIntValue
            dict_parameters_object[key] = valueObject

            if paramIsList != '':
                dictParamIsList[key] = paramIsList
                tempArr = []
                if paramIsList not in dictListWithParam:
                    dictListWithParam[paramIsList] = key
                else:

                    for val in dictListWithParam[paramIsList]:
                        if len(val) == 1:
                            tempArr.append(dictListWithParam[paramIsList])
                            break
                        else:
                            tempArr.append(val)

                    tempArr.append(key)
                    dictListWithParam[paramIsList] = tempArr
        df2 = pd.DataFrame()
        dfDesignValue = pd.read_excel(routeTemplate, sheet_name='Design Values')
        dfDesignValue.fillna("", inplace=True)

        for column in df:
            if column in arrParamTemp:
                df2[column] = df[column].values
            else:
                if dict_parameters[column] == 'Check "Design Values" sheet':
                    for i in range(len(df)):
                        df.at[i, column] = str(df.at[i, column]).replace('.0', '')
                    
                    for i in range(len(df2)):
                        for j in range(len(dfDesignValue)):
                            if str(df2.at[i, paramKey]).replace('.0', '') == str(dfDesignValue.at[j, paramKey]).replace('.0', ''):
                                valueTemplateIntValue = dfDesignValue.at[j, column]
                                if valueTemplateIntValue == '':
                                    valueTemplateIntValue = 0
                                elif valueTemplateIntValue != '' and valueTemplateIntValue != '#N/A':
                                    try:
                                        valueTemplateIntValue = int(dfDesignValue.at[j, column])
                                    except:
                                        valueTemplateIntValue = str(dfDesignValue.at[j, column]).replace('.0', '')
                                elif valueTemplateIntValue == '#N/A':
                                    valueTemplateIntValue = ''
                                df2.at[i, column] = valueTemplateIntValue
                else:
                    df2[column] = dict_parameters[column]

        dfComparation = pd.DataFrame()
        dfComparation = df == df2
        dfComparation['Apply_Plan'] = ''

        for i in range(len(dfComparation)):
            countFalse = 0
            for column in dfComparation:
                if column == 'Apply_Plan' and countFalse > 0 :
                    dfComparation.at[i, column] = 'YES'
                elif column == 'Apply_Plan' and countFalse == 0 :
                    dfComparation.at[i, column] = 'NO'
                elif column not in arrParamTemp:
                    if dfComparation.at[i, column] == False:
                        countFalse += 1

        df2['Apply_Plan'] = dfComparation['Apply_Plan']

        for i in range(len(df2)):
            countBlank = 0
            for column in df2:
                if column == 'Apply_Plan' and countBlank > 0 :
                    df2.at[i, column] = 'PEND_INFO' 
                elif column not in arrParamTemp:
                    if str(df2.at[i, column]) == 'nan':
                        countBlank += 1

        try:
            df2.to_excel(os.getcwd() + '\Audit_File' + templateName + '.xlsx', index=False)
        except:
            mbox.showerror('File is Open','Error. File Audit_File is Open. Please close file and try again.')
            sys.exit()
        
        try:
            dfComparation.to_excel(os.getcwd() + '\Comparation.xlsx', index=False)
        except:
            mbox.showerror('File is Open','Error. File Comparation.xlsx is Open. Please close file and try again.')
            sys.exit()

        # Put color to cells with different value
        wbComparation = load_workbook(os.getcwd() + '\Comparation.xlsx')
        wbTemplate = load_workbook(os.getcwd() + '\Audit_File' + templateName + '.xlsx')
        wsComparation = wbComparation.active
        wsTemplate = wbTemplate.active
        redFill = openpyxl.styles.colors.Color(rgb='00FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=redFill)
        dictParamColumn = {}
        col = 2
        while wsComparation.cell(row=1, column=col).value != None:
            activeCell = wsComparation.cell(row=1, column=col).value
            dictParamColumn[activeCell] = col
            col += 1
        
        col = 2
        while wsComparation.cell(row=1, column=col).value != None:
            activeCell = wsComparation.cell(row=1, column=col).value
            row_ = 2

            if activeCell not in arrParamTemp:
                while wsComparation.cell(row=row_, column=col).value != None:
                    if wsComparation.cell(row=row_, column=col).value == False:
                        wsTemplate.cell(row=row_, column=col).fill = my_fill
                        if activeCell in dictParamIsList:
                            for val in dictListWithParam[dictParamIsList[activeCell]]:
                                if len(val) == 1:
                                    val = dictListWithParam[dictParamIsList[activeCell]]
                                wsTemplate.cell(row=row_, column=dictParamColumn[val]).fill = my_fill
                    row_ += 1
            col += 1

        wbComparation.save(os.getcwd() + '\Comparation.xlsx')
        wbTemplate.save(os.getcwd() + '\Audit_File' + templateName + '.xlsx')
        os.remove(os.getcwd() + '\Comparation.xlsx')

        if option == 2:
            mbox.showinfo('Parameters to Audit', 'Excel Files Baseline_Query and Audit_File was created')
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Process to create xml files
if option == 3 or option == 4:
    templateName = routeTemplate.replace('Template', '').replace('.xlsx', '').split('_')
    templateName = '_' + templateName[-1]

    #Create a dataframe of comparation with Audit_File
    if option == 4:
        df2 = pd.read_excel(io= routeDB)
        excel_file = routeDB
        dfComparation = pd.DataFrame()
        wb = load_workbook(excel_file, data_only = True)
        ws = wb.active
        flagApplyPlan = False
        for column_ in range(1, ws.max_column + 1):
            columnName = ws.cell(row=1, column=column_).value
            if columnName == 'Apply_Plan':
                flagApplyPlan = True
            dfComparation[columnName] = ''

        if flagApplyPlan == False:
            dfComparation['Apply_Plan'] = ''

        for row_ in range(1, ws.max_row + 1):
            countFalse = 0
            if row_ != 1:
                for column_ in range(1, ws.max_column + 1):
                    color_in_hex = str(ws.cell(row=row_ , column=column_).fill.start_color.index)
                    cellValue = str(ws.cell(row=row_ , column=column_).value)
                    columnName = ws.cell(row=1, column=column_).value

                    if color_in_hex != '0' and color_in_hex != '00000000' and cellValue != '':
                        dfComparation.at[row_ - 2, columnName] = False
                        countFalse += 1
                    else:
                        dfComparation.at[row_ - 2, columnName] = True
                    
                    if str(ws.cell(row=1 , column=column_).value) == 'Apply_Plan' and countFalse > 0:
                        dfComparation.at[row_ - 2, columnName] = 'YES'
                    elif str(ws.cell(row=1 , column=column_).value) == 'Apply_Plan' and countFalse == 0:
                        dfComparation.at[row_ - 2, columnName] = 'NO'
                    elif flagApplyPlan == False and column_ == ws.max_column:
                        if countFalse > 0:
                            dfComparation.at[row_ - 1, columnName] = 'YES'
                        else:
                            dfComparation.at[row_ - 1, columnName] = 'NO'

        df_query = pd.read_excel(io= routeTemplate, sheet_name='Query')
        df_query.fillna("", inplace=True)

        for i in df_query.index.values:
            key = df_query.iloc[i, [6]].values[0]
            
            if key != '':
                paramKey = key
        
        dict_parameters = {}
        dict_parameters_object = {}
        dfParameterTemplate = pd.read_excel(routeTemplate, sheet_name='Parameter Template')
        dfParameterTemplate.fillna("", inplace=True)

        for i in dfParameterTemplate.index.values:
            key = dfParameterTemplate.iloc[i, [2]].values[0]
            valueTemplateIntValue = dfParameterTemplate.iloc[i, [4]].values[0]
            valueObject = dfParameterTemplate.iloc[i, [0]].values[0]
            dict_parameters[key] = valueTemplateIntValue
            dict_parameters_object[key] = valueObject

    #Start Process to create xml files
    today = str(date.today())
    now = datetime.now()
    date = today + 'T' + str(now.hour) + ':' + str(now.minute) + ':' + str(now.second)
    planName = 'Plan_Update_5G'
    rcInitial = ''
    arrDistName = []
    arrListAndDistName = []
    paramAndDistName = []
    arrRCParamKey = []
    moOpen = False
    isClosed = True
    listOpen = False
    moClosed = []
    for i in range(len(df2)):
        rcActual = df2.at[i, 'PLMN_id']
        if df2.at[i, 'Apply_Plan'] == 'YES':

            if rcActual + str(df2.at[i, paramKey]) not in arrRCParamKey:
                arrRCParamKey.append(rcActual + str(df2.at[i, paramKey]))

                if rcInitial != rcActual:
                    file = open(os.getcwd() + '/' + planName + '_' + rcActual + templateName + '_' + today + '.xml', 'w')
                    file.write('<?xml version="1.0" encoding="UTF-8"?>' + '\n')
                    file.write("<!DOCTYPE raml SYSTEM 'raml20.dtd'>" + '\n')
                    file.write('<raml version="2.0" xmlns="raml20.xsd">' + '\n')
                    file.write('  <cmData type="plan" scope="all" name="' + planName + '.xml">' + '\n')
                    file.write('    <header>' + '\n')
                    file.write('      <log dateTime="' + date + '" action="created" appInfo="PlanExporter"/>' + '\n')
                    file.write('    </header>' + '\n')
                    rcInitial = rcActual
                    isClosed = False

                for j in range(len(dfParameterTemplate)):
                    distName = str(dfParameterTemplate.at[j, 'distName'])
                    manageObj = str(dfParameterTemplate.at[j, 'MO PLAN'])
                    moVersion = str(dfParameterTemplate.at[j, 'MO VERSION'])
                    parameter = str(dfParameterTemplate.at[j, 'PARAMETER'])
                    isList = str(dfParameterTemplate.at[j, 'LIST'])
                    isConcat = str(dfParameterTemplate.at[j, 'CONCAT'])
                    valueParam = str(df2.at[i, parameter]).replace('.0', '')

                    if parameter == 'pucchF3DeltaF':
                        pausa = True
                    
                    nextRowJ = j + 1
                    if nextRowJ > dfParameterTemplate.shape[0] - 1:
                        flagEndRow = True
                        nextRowJ -= 1
                    else:
                        flagEndRow = False
                    
                    nextRowI = i + 1
                    if nextRowI > df2.shape[0] - 1:
                        nextRowI -= 1
                    
                    rcNext = df2.at[nextRowI, 'PLMN_id']
                    nextList = str(dfParameterTemplate.at[nextRowJ, 'LIST'])
                    nextDistName = str(dfParameterTemplate.at[nextRowJ, 'distName'])
                    nextManageObj = str(dfParameterTemplate.at[nextRowJ, 'MO PLAN'])

                    if dfComparation.at[i, parameter] == False:

                        for column in df2:
                            distName = distName.replace('$' + column, str(df2.at[i, column]))
                            if '$' not in distName:
                                break

                        if rcActual + distName not in arrDistName:
                            arrDistName.append(rcActual + distName)
                            file.write('    <managedObject class="' + manageObj + '" version="' + moVersion + '" distName="' + distName + '" operation="update">' + '\n')
                            moOpen = True
                        
                        if parameter + valueParam + rcActual + distName not in paramAndDistName:
                            tabulation = ''
                            if isList != '' and rcActual + isList + distName not in arrListAndDistName:
                                arrListAndDistName.append(rcActual + isList + distName)
                                if listOpen == True:
                                    file.write('		</item>' + '\n')
                                    file.write('	  </list>' + '\n')
                                    listOpen = False

                                file.write('	  <list name="' + isList + '">' + '\n')
                                file.write('        <item>' + '\n')
                                listOpen = True
                                tabulation = '    '
                            elif isList != '':
                                tabulation = '    '
                            elif isConcat != '':
                                parameter = isConcat + parameter

                            file.write(tabulation + '      <p name="' + parameter + '">' + valueParam + '</p>' + '\n')
                            paramAndDistName.append(parameter + valueParam + rcActual + distName)

                        if isList != nextList and nextList != '' and listOpen == True and rcNext + nextList + distName not in arrListAndDistName or flagEndRow == True:
                            file.write('		</item>' + '\n')
                            file.write('	  </list>' + '\n')
                            listOpen = False

                        for column in df2:
                            nextDistName = nextDistName.replace('$' + column, str(df2.at[i, column]))
                            if '$' not in nextDistName:
                                break

                        if manageObj != nextManageObj and rcNext + nextDistName not in arrDistName and rcActual + distName not in moClosed or flagEndRow == True:
                            file.write('    </managedObject>' + '\n')
                            moOpen = False
                            moClosed.append(rcActual + distName)

                    if isList != nextList and nextList != '' and listOpen == True:
                        file.write('		</item>' + '\n')
                        file.write('	  </list>' + '\n')
                        listOpen = False

                    if j == dfParameterTemplate.shape[0] - 1 and moOpen == True:
                        file.write('    </managedObject>' + '\n')
                        moOpen = False
                    elif manageObj != nextManageObj and moOpen == True:
                        file.write('    </managedObject>' + '\n')
                        moOpen = False

                nextRowI = i + 1
                if nextRowI > df2.shape[0] - 1:
                    nextRowI -= 1
                if rcActual != df2.at[nextRowI, 'PLMN_id'] or i == df2.shape[0] - 1:
                    file.write('  </cmData>' + '\n')
                    file.write('</raml>' + '\n')
                    file.close()
                    isClosed = True

        nextRowI = i + 1
        if nextRowI > df2.shape[0] - 1:
            nextRowI -= 1

        if (rcActual != df2.at[nextRowI, 'PLMN_id'] and isClosed == False) or (i == df2.shape[0] - 1 and isClosed == False):
            file.write('  </cmData>' + '\n')
            file.write('</raml>' + '\n')
            file.close()
            isClosed = True

    if option == 3:
        mbox.showinfo('Parameters to Audit', 'Excel Files and XML Files created')
    elif option == 4:
        mbox.showinfo('Parameters to Audit', 'XML Files created')